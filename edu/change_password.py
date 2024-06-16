try:
    from openpyxl import load_workbook
except ImportError:
    exit("请先 pip3 install openpyxl")
try:
    import requests
except ImportError:
    exit("请先 pip3 install requests")
import json
import os
import time
import multiprocessing

def process_account(row, header):
    try:
        req = requests.patch(f"https://eduzone.codemao.cn/edu/zone/students/{row[1]}/password", headers=header,
                             data="{}")
        if req.status_code == 200:
            pwd = json.loads(req.text).get("password")
            with open("pwd.txt", "a") as f:
                f.write(f"{row[0]}---{pwd}\n")
            return True
        else:
            if json.loads(req.text).get("error_code") == "10010":
                exit("Token异常")
            nnt = str(req.text)
            print(f"\n此账号请求失败，返回内容：{nnt}\n---------------")
    except Exception as e:
        exit(f"发生错误：{e}")
    return False


def process_accounts(sheet, header):
    for row in sheet.iter_rows(values_only=True):
        start_time = time.time()
        if process_account(row, header):
            ut = time.time() - start_time
            speed = int(60 / ut)
            print("\r" + f"正在改密中... | 账号 {row[0]} | 速度 {speed}个/分", end="")


if __name__ == '__main__':
    if os.path.exists("pwd.txt"):
        os.remove("pwd.txt")

    xlsx_file_path = input("输入.xlsx文件路径(按照格式): ")
    token = input("输入Token : ")
    header = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
        "Authorization": f"Bearer {token}"
    }
    wb = load_workbook(xlsx_file_path)
    sheet = wb.active
    all_account = sheet.max_row
    print(f"---------------\n总共{all_account}个账号待改密\n---------------")

    chunk_size = all_account // 8
    processes = []
    for i in range(8):
        start = i * chunk_size
        end = (i + 1) * chunk_size if i < 3 else all_account
        process = multiprocessing.Process(target=process_accounts, args=(sheet, header))
        process.start()
        processes.append(process)

    for process in processes:
        process.join()

    exit("改密成功，请使用此文件夹内的 login_txt.py 进行登录")
