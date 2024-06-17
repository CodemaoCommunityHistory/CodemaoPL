# @author:Wangs_official
from openpyxl import load_workbook
import requests
import json
import os
import time
import rich
import typing
import queue
import threading

# Constrants

WORKER_NUM = 2000
REPORT_INTERVAL = 10
RETRIES = 10


def login(username, pwd):
    for _ in range(RETRIES):
        try:
            url = "https://api.codemao.cn/tiger/v3/web/accounts/login"
            header = {
                "Accept": "*/*",
                "Accept-Encoding": "gzip, deflate, br",
                "Accept-Language": "zh-CN,zh;q=0.9",
                "Connection": "keep-alive",
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
            }
            post_text = json.dumps(
                {"pid": "65edCTyg", "identity": username, "password": pwd}
            )
            return requests.post(url, data=post_text, headers=header)
        except:
            pass
    return None




def check_worker(accounts_queue: queue.Queue, callback_queue: queue.Queue):
    while 1:
        data: list = accounts_queue.get()
        if data == "EXIT":
            return
        token = login(data[0], data[1])
        if token:
            callback_queue.put(token.json().get("auth", {}).get("token"))


def main():
    xlsx_path = input("XLSX file path: ")
    rich.print(f"[green]I[/green]: Worker number: {WORKER_NUM}")
    rich.print(f"[green]I[/green]: Report interval: {REPORT_INTERVAL}")
    rich.print(f"[green]I[/green]: Loading accounts")
    wb = load_workbook(xlsx_path)
    sheet = wb.active
    account_numbers = sheet.max_row
    rich.print(f"[green]I[/green]: Initing queue")
    mission_queue = queue.Queue()
    callback_queue = queue.Queue()
    for row in sheet.iter_rows(min_row=1, min_col=2, max_col=3, values_only=True):
        user, password = row
        mission_queue.put([user, password])
    rich.print("[green]I[/green]: Starting workers")
    for _ in range(WORKER_NUM):
        threading.Thread(
            target=check_worker, args=(mission_queue, callback_queue)
        ).start()
    rich.print("[green]I[/green]: Login started!")
    last_checked = 0
    while not mission_queue.empty():
        pending = mission_queue.qsize()
        checked = account_numbers - pending
        checks_per_sec = (checked - last_checked) / REPORT_INTERVAL
        eta_secs = pending / checks_per_sec
        failed=checked-callback_queue.qsize()-WORKER_NUM
        tokens=callback_queue.qsize()
        rich.print(
            f"[green]I[/green]: Pending logins: {pending} Tokens: [green]{tokens}[/green] Failed logins: [red]{failed}[/red] Fail rate: [red]{round(failed/checked*100, 2)}%[/red] ETA: [bold][cyan]{round(eta_secs//60)}[/cyan][/bold]m [bold][cyan]{round(eta_secs%60)}[/cyan][/bold]s"
        )
        time.sleep(REPORT_INTERVAL)
        last_checked = checked
    rich.print(
        f"[green]I[/green]: Pending logins: 0 Tokens: [green]{callback_queue.qsize()}[/green] Failed logins: [red]{account_numbers-callback_queue.qsize()}[/red]"
    )
    rich.print("[green]I[/green]: Writing tokens to tokens.txt")
    with open("tokens.txt", "w+") as f:
        tokens = []
        while not callback_queue.empty():
            tokens.append(callback_queue.get())
        f.write("\n".join(tokens))
    rich.print("[green]I[/green]: Done!")


if __name__ == "__main__":
    main()
